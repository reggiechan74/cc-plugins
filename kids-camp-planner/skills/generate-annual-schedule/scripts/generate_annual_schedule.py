#!/usr/bin/env python3
"""
Annual Schedule Generator

Reads summer assignments from a budget spreadsheet's Daily Schedule tab,
combines with PA days, winter break, and March break from a school calendar,
and generates a consolidated annual schedule as markdown and/or an xlsx tab.

Usage:
    # Basic: same provider for all children per period type
    python3 generate_annual_schedule.py \
      --xlsx examples/sample-budget.xlsx \
      --calendar skills/camp-planning/references/school-calendars/public-boards/tcdsb.md \
      --children "Emma,Liam" \
      --pa-day-provider "City of Toronto" \
      --break-provider "YMCA Cedar Glen" \
      --output-md camp-research/annual-schedule-2025-2026.md \
      --update-xlsx

    # Per-child, per-day overrides via JSON file
    python3 generate_annual_schedule.py \
      --xlsx examples/sample-budget.xlsx \
      --calendar skills/camp-planning/references/school-calendars/public-boards/tcdsb.md \
      --children "Emma,Liam" \
      --pa-day-provider "City of Toronto" \
      --break-provider "YMCA Cedar Glen" \
      --overrides overrides.json \
      --output-md camp-research/annual-schedule-2025-2026.md

Overrides JSON format (per-child, per-date provider assignments):
    {
      "2025-09-26": {"Emma": "Science Camp Toronto", "Liam": "City of Toronto"},
      "2025-12-22": {"Emma": "City of Toronto"},
      "2026-03-16": {"Emma": "Science Camp Toronto", "Liam": "Science Camp Toronto"}
    }
    Any child not listed for a date falls back to the period default
    (--pa-day-provider or --break-provider). Summer days from the
    spreadsheet can also be overridden.
"""

import argparse
import json
import re
import sys
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Dynamic child column layout constants
# ---------------------------------------------------------------------------

SHARED_PREFIX_COLS = 3  # Date, Day, Period/Week#
COLS_PER_CHILD = 6      # Camp Name, Before Care, Camp Fee, After Care, Lunch, Day Total


def calculate_total_cols(n_children):
    """Total columns = prefix + (N * child_block) + daily_total."""
    return SHARED_PREFIX_COLS + (n_children * COLS_PER_CHILD) + 1


def get_child_col_offsets(n_children):
    """Return 0-indexed column offsets for each child's camp name column."""
    return [SHARED_PREFIX_COLS + i * COLS_PER_CHILD for i in range(n_children)]


def validate_child_count(n):
    if n > 4:
        print(f"Error: Maximum 4 children supported. Got {n}.", file=sys.stderr)
        sys.exit(1)
    if n < 1:
        print("Error: At least 1 child required.", file=sys.stderr)
        sys.exit(1)


# ---------------------------------------------------------------------------
# Per-period rate differentiation
# ---------------------------------------------------------------------------

def resolve_period_rate(provider_rates, period_type):
    """Get rate for a specific period, falling back to summer if not available."""
    period_map = {
        "summer": "summer",
        "pa_day": "pa_day",
        "school_holiday": "pa_day",
        "winter_break": "break",
        "march_break": "break",
        "fall_break": "break",
    }
    rate_key = period_map.get(period_type, "summer")
    rate = provider_rates.get(rate_key)
    if rate is None:
        rate = provider_rates.get("summer", provider_rates)
    return rate


# ---------------------------------------------------------------------------
# Reading inputs
# ---------------------------------------------------------------------------

def read_provider_rates(xlsx_path):
    """Read provider daily rates from the Provider Comparison tab.

    Returns {provider_name: {daily, before, after, lunch, total}}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Provider Comparison"]
    rates = {}
    for row in ws.iter_rows(min_row=4, max_col=10, values_only=False):
        name = row[0].value  # col A
        if not name or name.startswith("8-Week"):
            break
        daily = row[2].value or 0   # col C - Daily Cost
        before = row[3].value or 0  # col D - Before Care/day
        after = row[4].value or 0   # col E - After Care/day
        lunch = row[5].value or 0   # col F - Lunch/day
        total = daily + before + after + lunch
        rates[name] = {
            "daily": daily,
            "before": before,
            "after": after,
            "lunch": lunch,
            "total": total,
        }
    wb.close()
    return rates


def read_summer_assignments(xlsx_path, children):
    """Read summer camp assignments from the Daily Schedule tab.

    Returns list of dicts: [{date, week, assignments: {child: camp_name}}]
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Daily Schedule"]
    days = []

    # Child camp columns: dynamically computed based on number of children
    child_cols = get_child_col_offsets(len(children))

    for row in ws.iter_rows(min_row=4, max_col=calculate_total_cols(len(children)), values_only=False):
        date_val = row[0].value  # col A
        if date_val is None or (isinstance(date_val, str) and date_val == "TOTAL"):
            break
        if isinstance(date_val, datetime):
            date_val = date_val.date()
        elif isinstance(date_val, str):
            date_val = datetime.strptime(date_val, "%Y-%m-%d").date()

        week_num = row[2].value  # col C
        assignments = {}
        for i, child in enumerate(children):
            if i < len(child_cols):
                camp = row[child_cols[i]].value or ""
                assignments[child] = camp

        days.append({
            "date": date_val,
            "week": week_num,
            "assignments": assignments,
        })
    wb.close()
    return days


def parse_calendar(calendar_path):
    """Parse school calendar markdown for PA days, winter break, March break.

    Returns {
        pa_days: [{date, purpose}],
        winter_break: {start_str, end_str},
        march_break: {start_str, end_str},
    }
    """
    with open(calendar_path) as f:
        content = f.read()

    # Parse PA Days table (matches "PA Days - Elementary" or just "PA Days")
    # Format: | # | Date | Day | Purpose |
    pa_days = []
    pa_section = re.search(
        r"### PA Days(?:\s*-\s*Elementary)?\s*\n\|.*\n\|[-|]+\n((?:\|.*\n)*)",
        content,
    )
    if pa_section:
        for line in pa_section.group(1).strip().split("\n"):
            cols = [c.strip() for c in line.split("|")]
            # cols[0] is empty (before first |), cols[1]=#, cols[2]=date, cols[3]=day, cols[4]=purpose
            if len(cols) >= 5 and cols[1].strip().isdigit():
                date_str = cols[2].strip()
                purpose = cols[4].strip() if len(cols) > 4 else ""
                pa_days.append({"date_str": date_str, "purpose": purpose})

    # Parse winter break (Christmas Break or Winter Break) from Holidays & Breaks table
    # Format: | Christmas Break | December 22, 2025 - January 2, 2026 | 8 |
    winter_break = None
    winter_match = re.search(
        r"\|\s*(?:Christmas|Winter) Break\s*\|\s*(.+?)\s*\|\s*\d+\s*\|",
        content,
    )
    if winter_match:
        date_range = winter_match.group(1).strip()
        parts = date_range.split(" - ")
        if len(parts) == 2:
            winter_break = {"start_str": parts[0].strip(), "end_str": parts[1].strip()}

    # Parse March Break from Holidays & Breaks table
    # Format: | Mid-Winter Break (March Break) | March 16-20, 2026 | 5 |
    march_break = None
    march_match = re.search(
        r"\|\s*Mid-Winter Break \(March Break\)\s*\|\s*(.+?)\s*\|\s*\d+\s*\|",
        content,
    )
    if march_match:
        date_range = march_match.group(1).strip()
        # Format: "March 16-20, 2026"
        m = re.match(r"(\w+)\s+(\d+)-(\d+),\s*(\d+)", date_range)
        if m:
            month_name, start_day, end_day, year = m.groups()
            march_break = {
                "start_str": f"{month_name} {start_day}, {year}",
                "end_str": f"{month_name} {end_day}, {year}",
            }

    # Parse single-day school holidays from Holidays & Breaks table
    # These are entries with Weekdays Off = 1 that aren't Christmas/March break
    school_holidays = []
    holidays_section = re.search(
        r"### Holidays & Breaks\s*\n\|.*\n\|[-|]+\n((?:\|.*\n)*)",
        content,
    )
    if holidays_section:
        for line in holidays_section.group(1).strip().split("\n"):
            cols = [c.strip() for c in line.split("|")]
            # cols: ['', name, date(s), weekdays_off, ...]
            if len(cols) < 4:
                continue
            name = cols[1].strip()
            date_str = cols[2].strip()
            weekdays_off_str = cols[3].strip()

            # Skip multi-day breaks (already handled as winter_break, march_break, fall_break)
            try:
                weekdays_off = int(weekdays_off_str)
            except ValueError:
                continue
            if weekdays_off != 1:
                continue

            # Skip if this is a break we already parse (Christmas, March, Winter)
            if any(skip in name.lower() for skip in ["christmas", "march break", "mid-winter", "winter break"]):
                continue

            school_holidays.append({
                "name": name,
                "date_str": date_str,
            })

    # Parse fall break (multi-day, typically private schools)
    # Match "Fall Break" or "Midterm break" with Weekdays Off > 1 in Oct/Nov
    fall_break = None
    fall_match = re.search(
        r"\|\s*(?:Fall Break|Midterm [Bb]reak)\s*\|\s*(.+?)\s*\|\s*(\d+)\s*\|",
        content,
    )
    if fall_match:
        date_range = fall_match.group(1).strip()
        weekdays_off = int(fall_match.group(2))
        if weekdays_off > 1:
            # Parse date range: "November 3-7, 2025" or "November 3, 2025 - November 7, 2025"
            # Short form: "Month DD-DD, YYYY"
            m = re.match(r"(\w+)\s+(\d+)-(\d+),\s*(\d+)", date_range)
            if m:
                month_name, start_day, end_day, year = m.groups()
                fall_break = {
                    "start_str": f"{month_name} {start_day}, {year}",
                    "end_str": f"{month_name} {end_day}, {year}",
                }
            else:
                # Long form: "Month DD, YYYY - Month DD, YYYY"
                parts = date_range.split(" - ")
                if len(parts) == 2:
                    fall_break = {
                        "start_str": parts[0].strip(),
                        "end_str": parts[1].strip(),
                    }

    return {
        "pa_days": pa_days,
        "winter_break": winter_break,
        "march_break": march_break,
        "school_holidays": school_holidays,
        "fall_break": fall_break,
    }


def parse_date_flexible(date_str):
    """Parse dates in various formats: 'September 26, 2025', 'March 16, 2026', etc."""
    formats = ["%B %d, %Y", "%b %d, %Y", "%Y-%m-%d"]
    for fmt in formats:
        try:
            return datetime.strptime(date_str.strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date: {date_str}")


def get_weekdays_between(start, end, exclude_dates=None):
    """Get all weekdays between start and end (inclusive), excluding specific dates."""
    exclude = set(exclude_dates or [])
    days = []
    current = start
    while current <= end:
        if current.weekday() < 5 and current not in exclude:
            days.append(current)
        current += timedelta(days=1)
    return days


def find_civic_holiday(year):
    """Find Ontario Civic Holiday (first Monday in August)."""
    aug_1 = date(year, 8, 1)
    days_until_monday = (7 - aug_1.weekday()) % 7
    if aug_1.weekday() == 0:
        return aug_1
    return aug_1 + timedelta(days=days_until_monday)


def get_summer_holidays(year):
    """Return Ontario summer statutory/civic holidays that most camps close for.

    Returns dict of {date: holiday_name}.
    """
    return {
        date(year, 7, 1): "Canada Day",
        find_civic_holiday(year): "Civic Holiday",
    }


# ---------------------------------------------------------------------------
# Building the annual schedule
# ---------------------------------------------------------------------------

def _resolve_assignments(d, children, default_provider, overrides):
    """Resolve per-child provider for a given date.

    For each child, checks the overrides dict for a date-specific provider.
    Falls back to default_provider if no override exists for that child+date.
    """
    d_iso = d.isoformat()
    day_overrides = overrides.get(d_iso, {})
    return {child: day_overrides.get(child, default_provider) for child in children}


def resolve_calendars(calendar_arg, children):
    """Resolve calendar argument(s) to per-child calendar paths.

    Accepts either:
      - A single path string (applies to all children)
      - A list of "ChildName:path" strings (per-child)

    Returns: {child_name: calendar_path}
    """
    if isinstance(calendar_arg, str):
        return {child: calendar_arg for child in children}

    calendars = {}
    for entry in calendar_arg:
        if ":" not in entry:
            raise ValueError(f"Multi-calendar entry must be 'ChildName:path', got: {entry}")
        child, path = entry.split(":", 1)
        calendars[child.strip()] = path.strip()

    for child in children:
        if child not in calendars:
            raise ValueError(f"No calendar specified for child: {child}")
    return calendars


def build_annual_days(summer_days, calendar_data, pa_provider, break_provider,
                      children, overrides=None, fall_break_provider=None):
    """Combine summer + non-summer into a unified list of annual days.

    Args:
        overrides: dict of {"YYYY-MM-DD": {"ChildName": "Provider"}}
            Per-child, per-date provider overrides. For summer days, overrides
            replace the spreadsheet assignment. For non-summer days, overrides
            replace the period default. Any child not listed for a date falls
            back to the period default (or spreadsheet value for summer).

    Returns list of dicts sorted by date:
    [{date, day_name, period, period_label, assignments: {child: camp}, notes}]
    """
    overrides = overrides or {}
    all_days = []
    seen_dates = set()  # Track dates to prevent double-counting

    # Summer days (from spreadsheet, with optional overrides)
    summer_holidays = get_summer_holidays(summer_days[0]["date"].year) if summer_days else {}
    for sd in summer_days:
        d = sd["date"]
        d_iso = d.isoformat()
        week = sd["week"]
        notes = f"Week {week}" if d.weekday() == 0 else ""
        # Flag summer statutory/civic holidays
        if d in summer_holidays:
            holiday_name = summer_holidays[d]
            notes = f"{holiday_name} -- VERIFY camp open" + (f" ({notes})" if notes else "")
        # Apply overrides on top of spreadsheet assignments
        assignments = dict(sd["assignments"])
        if d_iso in overrides:
            for child in children:
                if child in overrides[d_iso]:
                    assignments[child] = overrides[d_iso][child]
        all_days.append({
            "date": d,
            "day_name": d.strftime("%a"),
            "period": "summer",
            "period_label": f"Summer Wk {week}",
            "assignments": assignments,
            "notes": notes,
        })
        seen_dates.add(d)

    # PA days (skip if date already covered by another period)
    for i, pa in enumerate(calendar_data["pa_days"], 1):
        d = parse_date_flexible(pa["date_str"])
        if d in seen_dates:
            print(f"  Warning: PA Day {d.isoformat()} overlaps with another period, skipping", file=sys.stderr)
            continue
        purpose = pa.get("purpose", "")
        all_days.append({
            "date": d,
            "day_name": d.strftime("%a"),
            "period": "pa_day",
            "period_label": "PA Day",
            "assignments": _resolve_assignments(d, children, pa_provider, overrides),
            "notes": f"PA Day {i}" + (f" - {purpose}" if purpose else ""),
        })
        seen_dates.add(d)

    # Winter break (skip dates already covered)
    if calendar_data["winter_break"]:
        wb_start = parse_date_flexible(calendar_data["winter_break"]["start_str"])
        wb_end = parse_date_flexible(calendar_data["winter_break"]["end_str"])
        # Exclude stat holidays: Dec 25 (Christmas), Dec 26 (Boxing Day), Jan 1 (New Year's)
        stat_holidays = set()
        for year in [wb_start.year, wb_end.year]:
            stat_holidays.add(date(year, 12, 25))
            stat_holidays.add(date(year, 12, 26))
            stat_holidays.add(date(year, 1, 1))
        winter_days = get_weekdays_between(wb_start, wb_end, exclude_dates=stat_holidays)
        for d in winter_days:
            if d in seen_dates:
                continue
            notes = ""
            if d.month == 12 and d.day == 24:
                notes = "Christmas Eve -- VERIFY camp open"
            elif d.month == 12 and d.day == 31:
                notes = "New Year's Eve -- VERIFY camp open"
            all_days.append({
                "date": d,
                "day_name": d.strftime("%a"),
                "period": "winter_break",
                "period_label": "Winter Break",
                "assignments": _resolve_assignments(d, children, break_provider, overrides),
                "notes": notes,
            })
            seen_dates.add(d)

    # March break (skip dates already covered)
    if calendar_data["march_break"]:
        mb_start = parse_date_flexible(calendar_data["march_break"]["start_str"])
        mb_end = parse_date_flexible(calendar_data["march_break"]["end_str"])
        march_days = get_weekdays_between(mb_start, mb_end)
        for d in march_days:
            if d in seen_dates:
                continue
            all_days.append({
                "date": d,
                "day_name": d.strftime("%a"),
                "period": "march_break",
                "period_label": "March Break",
                "assignments": _resolve_assignments(d, children, break_provider, overrides),
                "notes": "",
            })
            seen_dates.add(d)

    # School holidays (single-day, use pa_provider for coverage)
    # Exclude: holidays during summer, during winter break, during March break
    # These are filtered by seen_dates set
    summer_start = summer_days[0]["date"] if summer_days else None
    summer_end = summer_days[-1]["date"] if summer_days else None
    for holiday in calendar_data.get("school_holidays", []):
        d = parse_date_flexible(holiday["date_str"])
        if d in seen_dates:
            continue
        # Skip holidays during summer period
        if summer_start and summer_end and summer_start <= d <= summer_end:
            continue
        # Skip Labour Day and other pre-school-year holidays
        if d.month in (6, 7, 8) or (d.month == 9 and d.day == 1):
            continue
        all_days.append({
            "date": d,
            "day_name": d.strftime("%a"),
            "period": "school_holiday",
            "period_label": "School Holiday",
            "assignments": _resolve_assignments(d, children, pa_provider, overrides),
            "notes": holiday["name"],
        })
        seen_dates.add(d)

    # Fall break (multi-day, use fall_break_provider or break_provider)
    fb_provider = fall_break_provider or break_provider
    if calendar_data.get("fall_break"):
        fb_start = parse_date_flexible(calendar_data["fall_break"]["start_str"])
        fb_end = parse_date_flexible(calendar_data["fall_break"]["end_str"])
        fall_days = get_weekdays_between(fb_start, fb_end)
        for d in fall_days:
            if d in seen_dates:
                continue
            all_days.append({
                "date": d,
                "day_name": d.strftime("%a"),
                "period": "fall_break",
                "period_label": "Fall Break",
                "assignments": _resolve_assignments(d, children, fb_provider, overrides),
                "notes": "",
            })
            seen_dates.add(d)

    all_days.sort(key=lambda x: x["date"])
    return all_days


def build_annual_days_multi(summer_days, per_child_calendars, pa_provider,
                            break_provider, children, overrides=None,
                            fall_break_provider=None):
    """Multi-calendar variant of build_annual_days.

    per_child_calendars: {child_name: calendar_data_dict}

    Merges all children's off-school dates. On days when only some children
    are off, others show "In school" with $0 cost.
    """
    overrides = overrides or {}
    all_days = []
    seen_dates = set()

    # Summer days (same for all children)
    summer_holidays = get_summer_holidays(summer_days[0]["date"].year) if summer_days else {}
    for sd in summer_days:
        d = sd["date"]
        d_iso = d.isoformat()
        week = sd["week"]
        notes = f"Week {week}" if d.weekday() == 0 else ""
        if d in summer_holidays:
            holiday_name = summer_holidays[d]
            notes = f"{holiday_name} -- VERIFY camp open" + (f" ({notes})" if notes else "")
        assignments = dict(sd["assignments"])
        if d_iso in overrides:
            for child in children:
                if child in overrides[d_iso]:
                    assignments[child] = overrides[d_iso][child]
        all_days.append({
            "date": d, "day_name": d.strftime("%a"), "period": "summer",
            "period_label": f"Summer Wk {week}", "assignments": assignments, "notes": notes,
        })
        seen_dates.add(d)

    # PA days — merge across all calendars
    all_pa_dates = {}
    for child, cal in per_child_calendars.items():
        for pa in cal.get("pa_days", []):
            d = parse_date_flexible(pa["date_str"])
            if d not in all_pa_dates:
                all_pa_dates[d] = {}
            all_pa_dates[d][child] = pa.get("purpose", "")

    pa_counter = 0
    for d in sorted(all_pa_dates.keys()):
        if d in seen_dates:
            continue
        pa_counter += 1
        children_off = all_pa_dates[d]
        assignments = {}
        for child in children:
            if child in children_off:
                d_iso = d.isoformat()
                day_overrides = overrides.get(d_iso, {})
                assignments[child] = day_overrides.get(child, pa_provider)
            else:
                assignments[child] = "In school"
        if len(children_off) < len(children):
            off_names = [c for c in children if c in children_off]
            notes = f"PA Day {pa_counter} ({', '.join(off_names)} off)"
        else:
            purpose = next(iter(children_off.values()), "")
            notes = f"PA Day {pa_counter}" + (f" - {purpose}" if purpose else "")
        all_days.append({
            "date": d, "day_name": d.strftime("%a"), "period": "pa_day",
            "period_label": "PA Day", "assignments": assignments, "notes": notes,
        })
        seen_dates.add(d)

    # School holidays — merge across all calendars
    all_holiday_dates = {}
    for child, cal in per_child_calendars.items():
        for h in cal.get("school_holidays", []):
            d = parse_date_flexible(h["date_str"])
            if d not in all_holiday_dates:
                all_holiday_dates[d] = {}
            all_holiday_dates[d][child] = h["name"]

    summer_start = summer_days[0]["date"] if summer_days else None
    summer_end = summer_days[-1]["date"] if summer_days else None
    for d in sorted(all_holiday_dates.keys()):
        if d in seen_dates:
            continue
        if summer_start and summer_end and summer_start <= d <= summer_end:
            continue
        if d.month in (6, 7, 8) or (d.month == 9 and d.day == 1):
            continue
        children_off = all_holiday_dates[d]
        assignments = {}
        for child in children:
            if child in children_off:
                d_iso = d.isoformat()
                day_overrides = overrides.get(d_iso, {})
                assignments[child] = day_overrides.get(child, pa_provider)
            else:
                assignments[child] = "In school"
        name = next(iter(children_off.values()))
        all_days.append({
            "date": d, "day_name": d.strftime("%a"), "period": "school_holiday",
            "period_label": "School Holiday", "assignments": assignments, "notes": name,
        })
        seen_dates.add(d)

    # Fall break — merge across calendars
    fb_provider = fall_break_provider or break_provider
    all_fall_dates = set()
    per_child_fall = {}
    for child, cal in per_child_calendars.items():
        if cal.get("fall_break"):
            fb_start = parse_date_flexible(cal["fall_break"]["start_str"])
            fb_end = parse_date_flexible(cal["fall_break"]["end_str"])
            child_dates = set(get_weekdays_between(fb_start, fb_end))
            per_child_fall[child] = child_dates
            all_fall_dates |= child_dates

    for d in sorted(all_fall_dates):
        if d in seen_dates:
            continue
        assignments = {}
        for child in children:
            if child in per_child_fall and d in per_child_fall[child]:
                d_iso = d.isoformat()
                day_overrides = overrides.get(d_iso, {})
                assignments[child] = day_overrides.get(child, fb_provider)
            else:
                assignments[child] = "In school"
        all_days.append({
            "date": d, "day_name": d.strftime("%a"), "period": "fall_break",
            "period_label": "Fall Break", "assignments": assignments, "notes": "",
        })
        seen_dates.add(d)

    # Winter break — union of all calendars
    all_winter_dates = set()
    per_child_winter = {}
    stat_holidays = set()
    for child, cal in per_child_calendars.items():
        if cal.get("winter_break"):
            wb_start = parse_date_flexible(cal["winter_break"]["start_str"])
            wb_end = parse_date_flexible(cal["winter_break"]["end_str"])
            for year in [wb_start.year, wb_end.year]:
                stat_holidays.add(date(year, 12, 25))
                stat_holidays.add(date(year, 12, 26))
                stat_holidays.add(date(year, 1, 1))
            child_dates = set(get_weekdays_between(wb_start, wb_end, exclude_dates=stat_holidays))
            per_child_winter[child] = child_dates
            all_winter_dates |= child_dates

    for d in sorted(all_winter_dates):
        if d in seen_dates:
            continue
        assignments = {}
        for child in children:
            if child in per_child_winter and d in per_child_winter[child]:
                d_iso = d.isoformat()
                day_overrides = overrides.get(d_iso, {})
                assignments[child] = day_overrides.get(child, break_provider)
            else:
                assignments[child] = "In school"
        notes = ""
        if d.month == 12 and d.day == 24:
            notes = "Christmas Eve -- VERIFY camp open"
        elif d.month == 12 and d.day == 31:
            notes = "New Year's Eve -- VERIFY camp open"
        all_days.append({
            "date": d, "day_name": d.strftime("%a"), "period": "winter_break",
            "period_label": "Winter Break", "assignments": assignments, "notes": notes,
        })
        seen_dates.add(d)

    # March break — union pattern
    all_march_dates = set()
    per_child_march = {}
    for child, cal in per_child_calendars.items():
        if cal.get("march_break"):
            mb_start = parse_date_flexible(cal["march_break"]["start_str"])
            mb_end = parse_date_flexible(cal["march_break"]["end_str"])
            child_dates = set(get_weekdays_between(mb_start, mb_end))
            per_child_march[child] = child_dates
            all_march_dates |= child_dates

    for d in sorted(all_march_dates):
        if d in seen_dates:
            continue
        assignments = {}
        for child in children:
            if child in per_child_march and d in per_child_march[child]:
                d_iso = d.isoformat()
                day_overrides = overrides.get(d_iso, {})
                assignments[child] = day_overrides.get(child, break_provider)
            else:
                assignments[child] = "In school"
        all_days.append({
            "date": d, "day_name": d.strftime("%a"), "period": "march_break",
            "period_label": "March Break", "assignments": assignments, "notes": "",
        })
        seen_dates.add(d)

    all_days.sort(key=lambda x: x["date"])
    return all_days


# ---------------------------------------------------------------------------
# Markdown rendering
# ---------------------------------------------------------------------------

def render_markdown(annual_days, provider_rates, children, render_context=None):
    """Render the annual schedule as markdown matching the sample format."""
    render_context = render_context or {}
    lines = []

    # Determine school year span
    dates = [d["date"] for d in annual_days]
    first_year = min(dates).year
    last_year = max(dates).year
    year_label = f"{first_year}-{last_year}" if first_year != last_year else str(first_year)

    lines.append(f"# {year_label} Annual Camp Schedule\n")
    lines.append(f"**Children**: {' & '.join(children)}")
    lines.append(f"**Generated**: {date.today().isoformat()}")
    lines.append("")
    lines.append("---")
    lines.append("")

    # Group days by period sections
    sections = _group_into_sections(annual_days)

    period_totals = {}  # {period_key: {child: cost, days: count}}

    for section in sections:
        period = section["period"]
        section_days = section["days"]
        title = section["title"]
        count = len(section_days)

        lines.append(f"## {title}")
        lines.append("")

        # Period description
        if period == "summer":
            week_nums = set()
            for d in section_days:
                week_nums.add(d["period_label"].replace("Summer Wk ", ""))
            total_weeks = len(week_nums) if week_nums else count // 5
            lines.append(f"{total_weeks} full weeks, {count} weekdays")
            lines.append("")
        elif period == "winter_break":
            lines.append(f"{count} camp-eligible weekdays (Dec 25-26 Christmas/Boxing Day and Jan 1 New Year's are stat holidays)")
            lines.append("")
        elif period == "march_break":
            lines.append(f"{count} weekdays")
            lines.append("")
        elif period == "fall_break":
            lines.append(f"{count} weekdays")
            lines.append("")

        # Table header
        child_headers = []
        for child in children:
            child_headers.extend([f"{child}'s Camp", f"{child} Cost"])
        header = "| Date | Day | " + " | ".join(child_headers) + " | Daily Total | Notes |"
        sep = "|------|-----|" + "|".join(["-------------|----------" for _ in children]) + "|-------------|-------|"
        lines.append(header)
        lines.append(sep)

        section_child_totals = {child: 0 for child in children}
        section_total = 0

        for day in section_days:
            d = day["date"]
            cells = [d.isoformat(), day["day_name"]]
            daily_total = 0
            for child in children:
                camp = day["assignments"].get(child, "")
                rate = provider_rates.get(camp, {}).get("total", 0)
                cells.extend([camp, f"${rate}"])
                section_child_totals[child] += rate
                daily_total += rate
            cells.extend([f"${daily_total}", day["notes"]])
            lines.append("| " + " | ".join(str(c) for c in cells) + " |")
            section_total += daily_total

        lines.append("")

        # Subtotals for multi-day periods
        period_key = section["period_key"]
        if period_key not in period_totals:
            period_totals[period_key] = {
                "period": period,
                "title": section["summary_title"],
                "days": 0,
                **{child: 0 for child in children},
                "combined": 0,
            }
        period_totals[period_key]["days"] += count
        for child in children:
            period_totals[period_key][child] += section_child_totals[child]
        period_totals[period_key]["combined"] += section_total

        if count > 1:
            parts = []
            for child in children:
                parts.append(f"{child} ${section_child_totals[child]:,}")
            parts.append(f"Combined ${section_total:,}")
            lines.append(f"**{section['subtotal_label']} subtotal**: " + " | ".join(parts))
            lines.append("")

        lines.append("---")
        lines.append("")

    # Annual summary table
    lines.append("## Annual Summary")
    lines.append("")
    child_headers = [f"{child} Total" for child in children]
    header = "| Period | Days | " + " | ".join(child_headers) + " | Combined |"
    sep = "|--------|------|" + "|".join(["----------" for _ in children]) + "|----------|"
    lines.append(header)
    lines.append(sep)

    grand_days = 0
    grand_child_totals = {child: 0 for child in children}
    grand_total = 0

    # Summarize by period type
    summary_order = ["summer", "pa_day", "school_holiday", "fall_break", "winter_break", "march_break"]
    summary_labels = {
        "summer": "Summer {year}",
        "pa_day": "PA Days ({count})",
        "school_holiday": "School Holidays ({count})",
        "fall_break": "Fall Break",
        "winter_break": "Winter Break",
        "march_break": "March Break",
    }

    for period_type in summary_order:
        # Aggregate all sections of this type
        days_count = 0
        child_sums = {child: 0 for child in children}
        combined = 0
        for pk, pt in period_totals.items():
            if pt["period"] == period_type:
                days_count += pt["days"]
                for child in children:
                    child_sums[child] += pt[child]
                combined += pt["combined"]
        if days_count == 0:
            continue

        label = summary_labels[period_type]
        if "{year}" in label:
            label = label.format(year=first_year)
        if "{count}" in label:
            label = label.format(count=days_count)

        child_cells = [f"${child_sums[child]:,}" for child in children]
        lines.append(f"| {label} | {days_count} | " + " | ".join(child_cells) + f" | ${combined:,} |")

        grand_days += days_count
        for child in children:
            grand_child_totals[child] += child_sums[child]
        grand_total += combined

    # Total row
    child_cells = [f"**${grand_child_totals[child]:,}**" for child in children]
    lines.append(f"| **Annual Total** | **{grand_days}** | " + " | ".join(child_cells) + f" | **${grand_total:,}** |")
    lines.append("")

    # Cost notes -- dynamically reference actual providers and rates
    lines.append("### Cost Notes")
    lines.append("")
    lines.append("- Summer costs include daily camp fee + before care + after care + lunch (from Provider Comparison rates)")

    # Use actual provider args and rates (passed in via render context)
    pa_prov = render_context.get("pa_provider", "")
    break_prov = render_context.get("break_provider", "")
    pa_rate = provider_rates.get(pa_prov, {}).get("total", 0)
    break_rate = provider_rates.get(break_prov, {}).get("total", 0)
    has_overrides = render_context.get("has_overrides", False)
    if has_overrides:
        lines.append(f"- PA day default: {pa_prov} (${pa_rate}/day); per-child overrides applied where specified")
        lines.append(f"- Break default: {break_prov} (${break_rate}/day); per-child overrides applied where specified")
    else:
        lines.append(f"- PA day costs use {pa_prov} rates (${pa_rate}/day all-in)")
        lines.append(f"- Winter and March Break use {break_prov} rates (${break_rate}/day all-in)")
    lines.append("- Costs shown are pre-discount; apply sibling/early-bird discounts to reduce totals")

    # List provider rates
    rate_parts = []
    for name, rates in sorted(provider_rates.items()):
        rate_parts.append(f"{name} ${rates['total']}/day")
    lines.append(f"- Daily rates: {', '.join(rate_parts)}")
    lines.append("- Non-summer rates are assumed equal to summer rates from Provider Comparison; confirm actual PA day and break program pricing with providers")
    lines.append("- Dates flagged VERIFY may be statutory/civic holidays when camps are closed")
    lines.append("")

    return "\n".join(lines)


def _group_into_sections(annual_days):
    """Group annual days into sections for markdown rendering.

    Summer is one big section. PA days and school holidays are individual sections.
    Winter break, March break, and fall break are each one section.
    """
    sections = []

    # Collect by period
    summer_days = [d for d in annual_days if d["period"] == "summer"]
    pa_days = [d for d in annual_days if d["period"] == "pa_day"]
    winter_days = [d for d in annual_days if d["period"] == "winter_break"]
    march_days = [d for d in annual_days if d["period"] == "march_break"]
    school_holiday_days = [d for d in annual_days if d["period"] == "school_holiday"]
    fall_break_days = [d for d in annual_days if d["period"] == "fall_break"]

    # Summer section
    if summer_days:
        first = summer_days[0]["date"]
        last = summer_days[-1]["date"]
        sections.append({
            "period": "summer",
            "period_key": "summer",
            "title": f"Summer {first.year} ({first.strftime('%b %d')} - {last.strftime('%b %d')})",
            "summary_title": f"Summer {first.year}",
            "subtotal_label": "Summer",
            "days": summer_days,
        })

    # PA days - each as its own section
    for pa in pa_days:
        d = pa["date"]
        sections.append({
            "period": "pa_day",
            "period_key": "pa_days",
            "title": f"PA Day: {d.strftime('%b %d, %Y')}",
            "summary_title": "PA Days",
            "subtotal_label": "PA Day",
            "days": [pa],
        })

    # Winter break
    if winter_days:
        first = winter_days[0]["date"]
        last = winter_days[-1]["date"]
        year_label = f"{first.year}-{last.year}" if first.year != last.year else str(first.year)
        sections.append({
            "period": "winter_break",
            "period_key": "winter_break",
            "title": f"Winter Break {year_label} ({first.strftime('%b %d')} - {last.strftime('%b %d')})",
            "summary_title": "Winter Break",
            "subtotal_label": "Winter Break",
            "days": winter_days,
        })

    # March break
    if march_days:
        first = march_days[0]["date"]
        last = march_days[-1]["date"]
        sections.append({
            "period": "march_break",
            "period_key": "march_break",
            "title": f"March Break {first.year} ({first.strftime('%b %d')}-{last.strftime('%b %d')})",
            "summary_title": "March Break",
            "subtotal_label": "March Break",
            "days": march_days,
        })

    # School holidays - each as its own section (like PA days)
    for sh in school_holiday_days:
        d = sh["date"]
        sections.append({
            "period": "school_holiday",
            "period_key": "school_holidays",
            "title": f"School Holiday: {sh['notes']} ({d.strftime('%b %d, %Y')})",
            "summary_title": "School Holidays",
            "subtotal_label": "School Holiday",
            "days": [sh],
        })

    # Fall break
    if fall_break_days:
        first = fall_break_days[0]["date"]
        last = fall_break_days[-1]["date"]
        sections.append({
            "period": "fall_break",
            "period_key": "fall_break",
            "title": f"Fall Break {first.year} ({first.strftime('%b %d')}-{last.strftime('%b %d')})",
            "summary_title": "Fall Break",
            "subtotal_label": "Fall Break",
            "days": fall_break_days,
        })

    # Sort sections chronologically by first day
    sections.sort(key=lambda s: s["days"][0]["date"])
    return sections


# ---------------------------------------------------------------------------
# XLSX update
# ---------------------------------------------------------------------------

def update_xlsx(xlsx_path, annual_days, children, provider_count=3):
    """Add or replace an 'Annual Schedule' tab in the xlsx.

    Uses dynamic column layout: 3 prefix + (N * 6 child block) + 1 daily total.
    VLOOKUP formulas reference Provider Comparison.
    """
    n_children = len(children)
    total_cols = calculate_total_cols(n_children)
    child_offsets = get_child_col_offsets(n_children)

    wb = openpyxl.load_workbook(xlsx_path)

    # Remove existing Annual Schedule tab if present
    if "Annual Schedule" in wb.sheetnames:
        del wb["Annual Schedule"]

    ws = wb.create_sheet("Annual Schedule")

    # VLOOKUP range reference
    vlookup_range = f"'Provider Comparison'!$A$4:$M${3 + provider_count}"

    # Row 1: Title
    ws["A1"] = "Annual Schedule"

    # Row 2: Instructions
    ws["A2"] = "Auto-generated from Daily Schedule + school calendar. Camp names reference Provider Comparison tab."

    # Row 3: Headers — dynamically generated for N children
    headers = ["Date", "Day", "Period"]
    for child in children:
        headers.extend([
            f"{child}'s Camp",
            f"{child} Before Care", f"{child} Camp",
            f"{child} After Care", f"{child} Lunch",
            f"{child} Day Total",
        ])
    headers.append("Daily Total")

    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Data rows start at row 4
    daily_total_col = total_cols  # last column

    for i, day in enumerate(annual_days):
        row = i + 4
        d = day["date"]

        # Col A: Date
        ws.cell(row=row, column=1, value=d)
        ws[f"A{row}"].number_format = "YYYY-MM-DD"

        # Col B: Day of week (formula)
        ws.cell(row=row, column=2, value=f'=TEXT(A{row},"ddd")')

        # Col C: Period label
        ws.cell(row=row, column=3, value=day["period_label"])

        # Per-child columns (6 cols each: Camp Name, Before Care, Camp Fee, After Care, Lunch, Day Total)
        child_day_total_cols = []
        for ci, child in enumerate(children):
            # 1-indexed column for camp name
            camp_col = child_offsets[ci] + 1  # convert 0-indexed offset to 1-indexed
            camp_letter = get_column_letter(camp_col)

            camp_name = day["assignments"].get(child, "")
            ws.cell(row=row, column=camp_col, value=camp_name)

            # Before Care (offset+1)
            before_col = camp_col + 1
            ws.cell(row=row, column=before_col,
                    value=f'=IF({camp_letter}{row}="","",VLOOKUP({camp_letter}{row},{vlookup_range},4,FALSE))')

            # Camp Fee (offset+2)
            fee_col = camp_col + 2
            ws.cell(row=row, column=fee_col,
                    value=f'=IF({camp_letter}{row}="","",VLOOKUP({camp_letter}{row},{vlookup_range},3,FALSE))')

            # After Care (offset+3)
            after_col = camp_col + 3
            ws.cell(row=row, column=after_col,
                    value=f'=IF({camp_letter}{row}="","",VLOOKUP({camp_letter}{row},{vlookup_range},5,FALSE))')

            # Lunch (offset+4)
            lunch_col = camp_col + 4
            ws.cell(row=row, column=lunch_col,
                    value=f'=IF({camp_letter}{row}="","",VLOOKUP({camp_letter}{row},{vlookup_range},6,FALSE))')

            # Child Day Total (offset+5) = SUM of before..lunch
            day_total_col = camp_col + 5
            before_letter = get_column_letter(before_col)
            lunch_letter = get_column_letter(lunch_col)
            ws.cell(row=row, column=day_total_col,
                    value=f"=SUM({before_letter}{row}:{lunch_letter}{row})")
            child_day_total_cols.append(day_total_col)

        # Daily Total = sum of all child day totals
        if len(child_day_total_cols) == 1:
            dt_letter = get_column_letter(child_day_total_cols[0])
            ws.cell(row=row, column=daily_total_col, value=f"={dt_letter}{row}")
        else:
            parts = "+".join(f"{get_column_letter(c)}{row}" for c in child_day_total_cols)
            ws.cell(row=row, column=daily_total_col, value=f"={parts}")

    # TOTAL row
    total_row = len(annual_days) + 4
    ws.cell(row=total_row, column=1, value="TOTAL")
    last_data_row = total_row - 1

    # SUM formulas for cost columns (all child cost columns + daily total)
    sum_cols = []
    for ci in range(n_children):
        camp_col = child_offsets[ci] + 1
        # Before care, camp fee, after care, lunch, child day total
        sum_cols.extend([camp_col + 1, camp_col + 2, camp_col + 3, camp_col + 4, camp_col + 5])
    sum_cols.append(daily_total_col)

    for col in sum_cols:
        col_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({col_letter}4:{col_letter}{last_data_row})")

    wb.save(xlsx_path)
    wb.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Annual Schedule Generator")
    parser.add_argument("--xlsx", required=True,
                        help="Path to budget spreadsheet (reads Provider Comparison + Daily Schedule)")
    parser.add_argument("--calendar", required=True,
                        help="Path to school calendar markdown file")
    parser.add_argument("--children", required=True,
                        help="Comma-separated children's names (e.g., 'Emma,Liam')")
    parser.add_argument("--pa-day-provider", default="City of Toronto",
                        help="Provider for PA day coverage")
    parser.add_argument("--break-provider", default="YMCA Cedar Glen",
                        help="Provider for winter and March break coverage")
    parser.add_argument("--fall-break-provider",
                        help="Provider for fall break coverage (default: same as --break-provider)")
    parser.add_argument("--overrides",
                        help="JSON file with per-child, per-date provider overrides. "
                             "Format: {\"YYYY-MM-DD\": {\"ChildName\": \"Provider\"}}")
    parser.add_argument("--output-md",
                        help="Path for output markdown file")
    parser.add_argument("--update-xlsx", action="store_true",
                        help="Add/replace Annual Schedule tab in the xlsx")

    args = parser.parse_args()
    children = [c.strip() for c in args.children.split(",")]

    validate_child_count(len(children))

    # Read inputs
    provider_rates = read_provider_rates(args.xlsx)
    summer_days = read_summer_assignments(args.xlsx, children)
    calendar_data = parse_calendar(args.calendar)

    # Load overrides
    overrides = {}
    if args.overrides:
        with open(args.overrides) as f:
            overrides = json.load(f)
        print(f"Loaded {len(overrides)} date overrides from {args.overrides}")

    # Validate all providers exist (defaults + every provider in overrides)
    all_providers = {args.pa_day_provider, args.break_provider}
    for date_key, child_map in overrides.items():
        for child_name, provider_name in child_map.items():
            all_providers.add(provider_name)
    missing = [p for p in all_providers if p not in provider_rates]
    if missing:
        print(f"Error: Provider(s) not found in Provider Comparison tab: {', '.join(sorted(missing))}", file=sys.stderr)
        print(f"Available providers: {', '.join(sorted(provider_rates.keys()))}", file=sys.stderr)
        sys.exit(1)

    # Build unified schedule
    annual_days = build_annual_days(
        summer_days, calendar_data,
        args.pa_day_provider, args.break_provider, children,
        overrides=overrides,
        fall_break_provider=args.fall_break_provider,
    )

    # Summary stats
    period_counts = {}
    for d in annual_days:
        p = d["period"]
        period_counts[p] = period_counts.get(p, 0) + 1
    total_days = len(annual_days)

    print(f"Annual schedule: {total_days} days")
    for p, c in sorted(period_counts.items()):
        print(f"  {p}: {c} days")

    # Generate markdown
    if args.output_md:
        ctx = {
            "pa_provider": args.pa_day_provider,
            "break_provider": args.break_provider,
            "has_overrides": bool(overrides),
        }
        md = render_markdown(annual_days, provider_rates, children, render_context=ctx)
        with open(args.output_md, "w") as f:
            f.write(md)
        print(f"Markdown written to {args.output_md}")

    # Update xlsx
    if args.update_xlsx:
        update_xlsx(args.xlsx, annual_days, children, provider_count=len(provider_rates))
        print(f"Annual Schedule tab added to {args.xlsx}")


if __name__ == "__main__":
    main()
