"""Tests for generate_annual_schedule.py."""

import json
import os
import sys
import tempfile
from datetime import date
from unittest.mock import patch

import openpyxl
import pytest

# Add script directory to path
sys.path.insert(0, os.path.dirname(__file__))
from generate_annual_schedule import (
    parse_calendar,
    parse_date_flexible,
    build_annual_days,
    build_annual_days_multi,
    resolve_calendars,
    get_weekdays_between,
    get_summer_holidays,
    find_civic_holiday,
    _group_into_sections,
    _read_rate_block,
    render_markdown,
    read_provider_rates,
    update_xlsx,
    calculate_total_cols,
    get_child_col_offsets,
    validate_child_count,
    resolve_period_rate,
    main,
)


# --- Fixtures: calendar markdown content ---

TCDSB_CALENDAR = """\
# Toronto Catholic District School Board (TCDSB)

## 2025-2026 School Year

### Key Dates
| Item | Date(s) |
|------|---------|
| First day of school | September 2, 2025 |
| Last day of classes (elementary) | June 25, 2026 |

### PA Days - Elementary
| # | Date | Day | Purpose |
|---|------|-----|---------|
| 1 | September 26, 2025 | Friday | Provincial Education Priorities |
| 2 | October 10, 2025 | Friday | Provincial Education Priorities |
| 3 | November 14, 2025 | Friday | Parent-Teacher Conferences |
| 4 | January 16, 2026 | Friday | Assessment, Evaluation and Reporting |
| 5 | February 13, 2026 | Friday | Parent-Teacher Conferences |
| 6 | June 5, 2026 | Friday | Assessment, Evaluation and Reporting |
| 7 | June 26, 2026 | Friday | Provincial Education Priorities |

### Holidays & Breaks
| Holiday/Break | Date(s) | Weekdays Off |
|---------------|---------|-------------|
| Labour Day | September 1, 2025 | 1 |
| Thanksgiving | October 13, 2025 | 1 |
| Christmas Break | December 22, 2025 - January 2, 2026 | 8 |
| Family Day | February 16, 2026 | 1 |
| Mid-Winter Break (March Break) | March 16-20, 2026 | 5 |
| Good Friday | April 3, 2026 | 1 |
| Easter Monday | April 6, 2026 | 1 |
| Victoria Day | May 18, 2026 | 1 |
"""

GIST_CALENDAR = """\
# German International School Toronto (GIST)

## 2025-2026 School Year

### PA Days
| # | Date | Day | Purpose |
|---|------|-----|---------|
| 1 | October 24, 2025 | Friday | Kindergarten only |
| 2 | November 21, 2025 | Friday | Learning Development Meetings |
| 3 | January 30, 2026 | Friday | No classes |
| 4 | May 29, 2026 | Friday | No classes |

### Holidays & Breaks
| Holiday/Break | Date(s) | Weekdays Off | TDSB Equivalent |
|---------------|---------|-------------|-----------------|
| Labour Day | September 1, 2025 | 1 | Same |
| National Day for Truth & Reconciliation | September 30, 2025 | 1 | Same |
| Thanksgiving | October 13, 2025 | 1 | Same |
| Fall Break | November 3-7, 2025 | 5 | None - TDSB has no fall break |
| Remembrance Day | November 11, 2025 | 1 | Same |
| Christmas Break | December 22, 2025 - January 2, 2026 | 8 | Same |
| Family Day | February 16, 2026 | 1 | Same |
| March Break | March 16-27, 2026 | 10 | March 16-20 only (5 days) |
| Good Friday | April 3, 2026 | 1 | Same |
| Easter Monday | April 6, 2026 | 1 | Same |
| Victoria Day | May 18, 2026 | 1 | Same |
"""

# Calendar with no fall break (public board)
TDSB_CALENDAR = """\
# Toronto District School Board (TDSB)

## 2025-2026 School Year

### PA Days - Elementary
| # | Date | Day | Purpose |
|---|------|-----|---------|
| 1 | September 26, 2025 | Friday | Professional Development |
| 2 | October 10, 2025 | Friday | Professional Development |
| 3 | November 14, 2025 | Friday | Parent Teacher Conferences |
| 4 | January 16, 2026 | Friday | Assessment and Reporting |
| 5 | February 13, 2026 | Friday | Parent Teacher Conferences |
| 6 | June 5, 2026 | Friday | Assessment and Reporting |
| 7 | June 26, 2026 | Friday | Professional Development |

### Holidays & Breaks
| Holiday/Break | Date(s) | Weekdays Off |
|---------------|---------|-------------|
| Labour Day | September 1, 2025 | 1 |
| Thanksgiving | October 13, 2025 | 1 |
| Winter Break | December 22, 2025 - January 2, 2026 | 8 |
| Family Day | February 16, 2026 | 1 |
| Mid-Winter Break (March Break) | March 16-20, 2026 | 5 |
| Good Friday | April 3, 2026 | 1 |
| Easter Monday | April 6, 2026 | 1 |
| Victoria Day | May 18, 2026 | 1 |
"""


def _write_temp_calendar(content):
    """Write calendar content to a temp file and return its path."""
    f = tempfile.NamedTemporaryFile(mode="w", suffix=".md", delete=False)
    f.write(content)
    f.close()
    return f.name


# --- Existing functionality tests (regression) ---

class TestParseDateFlexible:
    def test_full_month_name(self):
        assert parse_date_flexible("September 26, 2025") == date(2025, 9, 26)

    def test_iso_format(self):
        assert parse_date_flexible("2025-09-26") == date(2025, 9, 26)

    def test_invalid_date_raises(self):
        with pytest.raises(ValueError):
            parse_date_flexible("not a date")


class TestGetWeekdaysBetween:
    def test_single_week(self):
        days = get_weekdays_between(date(2025, 7, 7), date(2025, 7, 11))
        assert len(days) == 5
        assert all(d.weekday() < 5 for d in days)

    def test_with_exclusion(self):
        days = get_weekdays_between(
            date(2025, 7, 7), date(2025, 7, 11),
            exclude_dates={date(2025, 7, 9)}
        )
        assert len(days) == 4
        assert date(2025, 7, 9) not in days


class TestFindCivicHoliday:
    def test_2025(self):
        assert find_civic_holiday(2025) == date(2025, 8, 4)

    def test_2026(self):
        assert find_civic_holiday(2026) == date(2026, 8, 3)


class TestGetSummerHolidays:
    def test_returns_canada_day_and_civic(self):
        holidays = get_summer_holidays(2025)
        assert date(2025, 7, 1) in holidays
        assert holidays[date(2025, 7, 1)] == "Canada Day"
        civic = find_civic_holiday(2025)
        assert civic in holidays


class TestParseCalendarExisting:
    """Regression tests for existing parse_calendar behavior."""

    def test_tcdsb_pa_days(self):
        path = _write_temp_calendar(TCDSB_CALENDAR)
        try:
            result = parse_calendar(path)
            assert len(result["pa_days"]) == 7
            assert result["pa_days"][0]["date_str"] == "September 26, 2025"
        finally:
            os.unlink(path)

    def test_tcdsb_winter_break(self):
        path = _write_temp_calendar(TCDSB_CALENDAR)
        try:
            result = parse_calendar(path)
            assert result["winter_break"]["start_str"] == "December 22, 2025"
            assert result["winter_break"]["end_str"] == "January 2, 2026"
        finally:
            os.unlink(path)

    def test_tcdsb_march_break(self):
        path = _write_temp_calendar(TCDSB_CALENDAR)
        try:
            result = parse_calendar(path)
            assert result["march_break"]["start_str"] == "March 16, 2026"
            assert result["march_break"]["end_str"] == "March 20, 2026"
        finally:
            os.unlink(path)

    def test_tdsb_winter_break_label(self):
        """TDSB uses 'Winter Break' instead of 'Christmas Break'."""
        path = _write_temp_calendar(TDSB_CALENDAR)
        try:
            result = parse_calendar(path)
            assert result["winter_break"] is not None
            assert result["winter_break"]["start_str"] == "December 22, 2025"
        finally:
            os.unlink(path)


class TestParseCalendarSchoolHolidays:
    """Tests for new school_holidays parsing."""

    def test_tcdsb_school_holidays_parsed(self):
        path = _write_temp_calendar(TCDSB_CALENDAR)
        try:
            result = parse_calendar(path)
            assert "school_holidays" in result
            holidays = result["school_holidays"]
            names = [h["name"] for h in holidays]
            assert "Thanksgiving" in names
            assert "Family Day" in names
            assert "Good Friday" in names
            assert "Easter Monday" in names
            assert "Victoria Day" in names
        finally:
            os.unlink(path)

    def test_tcdsb_holiday_dates_correct(self):
        path = _write_temp_calendar(TCDSB_CALENDAR)
        try:
            result = parse_calendar(path)
            holiday_map = {h["name"]: h["date_str"] for h in result["school_holidays"]}
            assert holiday_map["Thanksgiving"] == "October 13, 2025"
            assert holiday_map["Family Day"] == "February 16, 2026"
            assert holiday_map["Good Friday"] == "April 3, 2026"
            assert holiday_map["Easter Monday"] == "April 6, 2026"
            assert holiday_map["Victoria Day"] == "May 18, 2026"
        finally:
            os.unlink(path)

    def test_gist_holidays_parsed(self):
        path = _write_temp_calendar(GIST_CALENDAR)
        try:
            result = parse_calendar(path)
            names = [h["name"] for h in result["school_holidays"]]
            assert "Thanksgiving" in names
            assert "Good Friday" in names
            assert "Easter Monday" in names
            assert "Victoria Day" in names
            assert "Remembrance Day" in names
        finally:
            os.unlink(path)

    def test_holidays_exclude_multi_day_breaks(self):
        """Multi-day breaks (Christmas, March) should NOT appear in school_holidays."""
        path = _write_temp_calendar(TCDSB_CALENDAR)
        try:
            result = parse_calendar(path)
            names = [h["name"] for h in result["school_holidays"]]
            assert "Christmas Break" not in names
            assert "Mid-Winter Break (March Break)" not in names
        finally:
            os.unlink(path)


class TestParseCalendarFallBreak:
    """Tests for fall_break parsing."""

    def test_gist_fall_break_parsed(self):
        path = _write_temp_calendar(GIST_CALENDAR)
        try:
            result = parse_calendar(path)
            assert "fall_break" in result
            assert result["fall_break"] is not None
            assert result["fall_break"]["start_str"] == "November 3, 2025"
            assert result["fall_break"]["end_str"] == "November 7, 2025"
        finally:
            os.unlink(path)

    def test_tdsb_no_fall_break(self):
        path = _write_temp_calendar(TDSB_CALENDAR)
        try:
            result = parse_calendar(path)
            assert result["fall_break"] is None
        finally:
            os.unlink(path)

    def test_tcdsb_no_fall_break(self):
        path = _write_temp_calendar(TCDSB_CALENDAR)
        try:
            result = parse_calendar(path)
            assert result["fall_break"] is None
        finally:
            os.unlink(path)


class TestBuildAnnualDaysSchoolHolidays:
    """Tests for school holidays in build_annual_days."""

    def _make_summer_days(self):
        """Create minimal summer days fixture."""
        return [
            {"date": date(2025, 6, 30), "week": 1, "assignments": {"Emma": "YMCA", "Liam": "YMCA"}},
            {"date": date(2025, 7, 1), "week": 1, "assignments": {"Emma": "YMCA", "Liam": "YMCA"}},
        ]

    def test_school_holidays_included(self):
        cal = parse_calendar(_write_temp_calendar(TCDSB_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma", "Liam"])
        periods = {d["period"] for d in days}
        assert "school_holiday" in periods

    def test_school_holiday_uses_pa_provider(self):
        cal = parse_calendar(_write_temp_calendar(TCDSB_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma", "Liam"])
        holiday_days = [d for d in days if d["period"] == "school_holiday"]
        for d in holiday_days:
            for child in ["Emma", "Liam"]:
                assert d["assignments"][child] == "City of Toronto"

    def test_school_holidays_not_during_summer(self):
        """Canada Day (Jul 1) should NOT appear as a school_holiday."""
        cal = parse_calendar(_write_temp_calendar(TCDSB_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma", "Liam"])
        holiday_days = [d for d in days if d["period"] == "school_holiday"]
        holiday_dates = {d["date"] for d in holiday_days}
        assert date(2025, 7, 1) not in holiday_dates  # Canada Day is summer

    def test_school_holidays_not_during_breaks(self):
        """Holidays during winter/March break should not be double-counted."""
        cal = parse_calendar(_write_temp_calendar(TCDSB_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma", "Liam"])
        # Count total days - should not have duplicates
        all_dates = [d["date"] for d in days]
        assert len(all_dates) == len(set(all_dates)), "Duplicate dates found!"

    def test_tcdsb_school_holiday_count(self):
        """TCDSB should have 5 school holidays: Thanksgiving, Family Day, Good Friday, Easter Monday, Victoria Day."""
        cal = parse_calendar(_write_temp_calendar(TCDSB_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma", "Liam"])
        school_holidays = [d for d in days if d["period"] == "school_holiday"]
        assert len(school_holidays) == 5

    def test_labour_day_excluded(self):
        """Labour Day (Sep 1) is before school starts - excluded from coverage."""
        cal = parse_calendar(_write_temp_calendar(TCDSB_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma", "Liam"])
        holiday_days = [d for d in days if d["period"] == "school_holiday"]
        holiday_dates = {d["date"] for d in holiday_days}
        assert date(2025, 9, 1) not in holiday_dates


class TestBuildAnnualDaysFallBreak:
    """Tests for fall break in build_annual_days."""

    def _make_summer_days(self):
        return [
            {"date": date(2025, 6, 30), "week": 1, "assignments": {"Emma": "YMCA"}},
        ]

    def test_gist_fall_break_included(self):
        cal = parse_calendar(_write_temp_calendar(GIST_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma"])
        periods = {d["period"] for d in days}
        assert "fall_break" in periods

    def test_fall_break_uses_break_provider_default(self):
        cal = parse_calendar(_write_temp_calendar(GIST_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma"])
        fall_days = [d for d in days if d["period"] == "fall_break"]
        for d in fall_days:
            assert d["assignments"]["Emma"] == "YMCA"

    def test_fall_break_uses_custom_provider(self):
        cal = parse_calendar(_write_temp_calendar(GIST_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(
            summer, cal, "City of Toronto", "YMCA", ["Emma"],
            fall_break_provider="Science Camp"
        )
        fall_days = [d for d in days if d["period"] == "fall_break"]
        for d in fall_days:
            assert d["assignments"]["Emma"] == "Science Camp"

    def test_fall_break_day_count(self):
        """GIST Nov 3-7 = 5 weekdays."""
        cal = parse_calendar(_write_temp_calendar(GIST_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma"])
        fall_days = [d for d in days if d["period"] == "fall_break"]
        assert len(fall_days) == 5

    def test_tdsb_no_fall_break_days(self):
        cal = parse_calendar(_write_temp_calendar(TDSB_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma"])
        fall_days = [d for d in days if d["period"] == "fall_break"]
        assert len(fall_days) == 0

    def test_no_double_counting_with_pa_days(self):
        """Fall break days should not overlap with PA days."""
        cal = parse_calendar(_write_temp_calendar(GIST_CALENDAR))
        summer = self._make_summer_days()
        days = build_annual_days(summer, cal, "City of Toronto", "YMCA", ["Emma"])
        all_dates = [d["date"] for d in days]
        assert len(all_dates) == len(set(all_dates))


class TestGroupIntoSectionsNewPeriods:
    """Tests for new period types in _group_into_sections."""

    def test_school_holidays_grouped_individually(self):
        """Each school holiday should be its own section (like PA days)."""
        days = [
            {"date": date(2025, 10, 13), "day_name": "Mon", "period": "school_holiday",
             "period_label": "School Holiday", "assignments": {"Emma": "CoT"}, "notes": "Thanksgiving"},
            {"date": date(2026, 2, 16), "day_name": "Mon", "period": "school_holiday",
             "period_label": "School Holiday", "assignments": {"Emma": "CoT"}, "notes": "Family Day"},
        ]
        sections = _group_into_sections(days)
        assert len(sections) == 2
        assert sections[0]["period"] == "school_holiday"
        assert "Thanksgiving" in sections[0]["title"]

    def test_fall_break_single_section(self):
        """Fall break should be one section (like winter/March break)."""
        days = [
            {"date": date(2025, 11, d), "day_name": "Mon", "period": "fall_break",
             "period_label": "Fall Break", "assignments": {"Emma": "YMCA"}, "notes": ""}
            for d in range(3, 8)
        ]
        sections = _group_into_sections(days)
        assert len(sections) == 1
        assert sections[0]["period"] == "fall_break"
        assert "Fall Break" in sections[0]["title"]

    def test_mixed_periods_sorted_chronologically(self):
        days = [
            {"date": date(2025, 10, 13), "day_name": "Mon", "period": "school_holiday",
             "period_label": "School Holiday", "assignments": {"E": "C"}, "notes": "Thanksgiving"},
            {"date": date(2025, 11, 3), "day_name": "Mon", "period": "fall_break",
             "period_label": "Fall Break", "assignments": {"E": "C"}, "notes": ""},
            {"date": date(2025, 9, 26), "day_name": "Fri", "period": "pa_day",
             "period_label": "PA Day", "assignments": {"E": "C"}, "notes": "PA Day 1"},
        ]
        sections = _group_into_sections(days)
        dates = [s["days"][0]["date"] for s in sections]
        assert dates == sorted(dates)


class TestRenderMarkdownNewPeriods:
    """Tests for new period types in render_markdown and annual summary."""

    def test_summary_includes_school_holidays(self):
        days = [
            {"date": date(2025, 10, 13), "day_name": "Mon", "period": "school_holiday",
             "period_label": "School Holiday", "assignments": {"Emma": "City of Toronto"},
             "notes": "Thanksgiving"},
        ]
        rates = {"City of Toronto": {"daily": 50, "before": 5, "after": 5, "lunch": 2, "total": 62}}
        md = render_markdown(days, rates, ["Emma"])
        assert "School Holidays" in md
        assert "Thanksgiving" in md

    def test_summary_includes_fall_break(self):
        days = [
            {"date": date(2025, 11, d), "day_name": "Mon", "period": "fall_break",
             "period_label": "Fall Break", "assignments": {"Emma": "YMCA"},
             "notes": ""}
            for d in range(3, 8)
        ]
        rates = {"YMCA": {"daily": 70, "before": 8, "after": 8, "lunch": 1, "total": 87}}
        md = render_markdown(days, rates, ["Emma"])
        assert "Fall Break" in md


class TestUpdateXlsxNewPeriods:
    """Verify xlsx includes new period types."""

    def test_xlsx_contains_school_holiday_rows(self):
        import tempfile, shutil
        # Copy sample budget xlsx to temp
        src = os.path.join(os.path.dirname(__file__), "..", "..", "..", "examples", "sample-budget.xlsx")
        if not os.path.exists(src):
            pytest.skip("sample-budget.xlsx not found")
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copy2(src, tmp.name)
        try:
            from generate_annual_schedule import update_xlsx
            days = [
                {"date": date(2025, 10, 13), "day_name": "Mon", "period": "school_holiday",
                 "period_label": "School Holiday", "assignments": {"Emma": "City of Toronto", "Liam": "City of Toronto"},
                 "notes": "Thanksgiving"},
                {"date": date(2025, 11, 3), "day_name": "Mon", "period": "fall_break",
                 "period_label": "Fall Break", "assignments": {"Emma": "YMCA Cedar Glen", "Liam": "YMCA Cedar Glen"},
                 "notes": ""},
            ]
            update_xlsx(tmp.name, days, ["Emma", "Liam"])
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            ws = wb["Annual Schedule"]
            assert ws["C4"].value == "School Holiday"
            assert ws["C5"].value == "Fall Break"
            wb.close()
        finally:
            os.unlink(tmp.name)


class TestMultiCalendarCLI:
    """Tests for multi-calendar CLI arg parsing."""

    def test_single_calendar_backward_compat(self):
        """Single --calendar path applies to all children."""
        path = _write_temp_calendar(TCDSB_CALENDAR)
        try:
            calendars = resolve_calendars(path, ["Emma", "Liam"])
            assert calendars["Emma"] == path
            assert calendars["Liam"] == path
        finally:
            os.unlink(path)

    def test_multi_calendar_per_child(self):
        path1 = _write_temp_calendar(TDSB_CALENDAR)
        path2 = _write_temp_calendar(GIST_CALENDAR)
        try:
            calendars = resolve_calendars(
                [f"Emma:{path1}", f"Liam:{path2}"],
                ["Emma", "Liam"]
            )
            assert calendars["Emma"] == path1
            assert calendars["Liam"] == path2
        finally:
            os.unlink(path1)
            os.unlink(path2)


class TestMultiCalendarMerge:
    """Tests for merging calendars from different schools."""

    def test_merged_includes_both_pa_days(self):
        """TDSB and GIST PA days are different — merged calendar has both sets."""
        tdsb_path = _write_temp_calendar(TDSB_CALENDAR)
        gist_path = _write_temp_calendar(GIST_CALENDAR)
        try:
            cal_emma = parse_calendar(tdsb_path)
            cal_liam = parse_calendar(gist_path)
            summer = [{"date": date(2025, 6, 30), "week": 1,
                       "assignments": {"Emma": "YMCA", "Liam": "YMCA"}}]
            days = build_annual_days_multi(
                summer,
                {"Emma": cal_emma, "Liam": cal_liam},
                "City of Toronto", "YMCA",
                ["Emma", "Liam"],
            )
            pa_days = [d for d in days if d["period"] == "pa_day"]
            pa_dates = {d["date"] for d in pa_days}
            # TDSB PA: Sep 26, Oct 10, Nov 14, Jan 16, Feb 13, Jun 5, Jun 26
            # GIST PA: Oct 24, Nov 21, Jan 30, May 29
            # Merged unique: 11 PA days
            assert date(2025, 9, 26) in pa_dates  # TDSB only
            assert date(2025, 10, 24) in pa_dates  # GIST only
        finally:
            os.unlink(tdsb_path)
            os.unlink(gist_path)

    def test_in_school_child_marked(self):
        """When only one child is off, the other shows 'In school' with $0."""
        tdsb_path = _write_temp_calendar(TDSB_CALENDAR)
        gist_path = _write_temp_calendar(GIST_CALENDAR)
        try:
            cal_emma = parse_calendar(tdsb_path)
            cal_liam = parse_calendar(gist_path)
            summer = [{"date": date(2025, 6, 30), "week": 1,
                       "assignments": {"Emma": "YMCA", "Liam": "YMCA"}}]
            days = build_annual_days_multi(
                summer,
                {"Emma": cal_emma, "Liam": cal_liam},
                "City of Toronto", "YMCA",
                ["Emma", "Liam"],
            )
            # Oct 24 is GIST PA day only — Emma is in school
            oct24 = [d for d in days if d["date"] == date(2025, 10, 24)]
            assert len(oct24) == 1
            assert oct24[0]["assignments"]["Liam"] == "City of Toronto"
            assert oct24[0]["assignments"]["Emma"] == "In school"
        finally:
            os.unlink(tdsb_path)
            os.unlink(gist_path)


class TestDynamicChildColumns:
    """Tests for dynamic child column layout."""

    def test_column_formula_2_children(self):
        """2 children: 3 + 12 + 1 = 16 cols (backward compat)."""
        assert calculate_total_cols(2) == 16

    def test_column_formula_3_children(self):
        """3 children: 3 + 18 + 1 = 22 cols."""
        assert calculate_total_cols(3) == 22

    def test_column_formula_4_children(self):
        """4 children: 3 + 24 + 1 = 28 cols."""
        assert calculate_total_cols(4) == 28

    def test_child_col_offsets_2(self):
        offsets = get_child_col_offsets(2)
        assert offsets == [3, 9]  # 0-indexed: D=3, J=9

    def test_child_col_offsets_3(self):
        offsets = get_child_col_offsets(3)
        assert offsets == [3, 9, 15]

    def test_child_col_offsets_4(self):
        offsets = get_child_col_offsets(4)
        assert offsets == [3, 9, 15, 21]

    def test_max_4_children_enforced(self):
        with pytest.raises(SystemExit):
            validate_child_count(5)

    def test_1_child_supported(self):
        assert calculate_total_cols(1) == 10  # 3 + 6 + 1


class TestPerPeriodRates:
    """Tests for per-period rate differentiation."""

    def test_fallback_to_summer_rates(self):
        """When PA Day/Break columns are empty, fall back to summer rates."""
        rates = {"summer": {"daily": 60, "before": 10, "after": 10, "lunch": 7, "total": 87},
                 "pa_day": None, "break": None}
        resolved = resolve_period_rate(rates, "pa_day")
        assert resolved == rates["summer"]

    def test_period_specific_rate(self):
        rates = {"summer": {"daily": 60, "before": 10, "after": 10, "lunch": 7, "total": 87},
                 "pa_day": {"daily": 45, "before": 8, "after": 8, "lunch": 5, "total": 66},
                 "break": None}
        resolved = resolve_period_rate(rates, "pa_day")
        assert resolved["total"] == 66

    def test_break_rate_for_winter(self):
        rates = {"summer": {"daily": 60, "before": 10, "after": 10, "lunch": 7, "total": 87},
                 "pa_day": None,
                 "break": {"daily": 50, "before": 8, "after": 8, "lunch": 5, "total": 71}}
        resolved = resolve_period_rate(rates, "winter_break")
        assert resolved["total"] == 71


class TestMultiCalendarCLIIntegration:
    """Tests that --calendar works with action='append' in main()."""

    def test_argparse_accepts_multiple_calendars(self):
        """--calendar should accept multiple values via action='append'."""
        from generate_annual_schedule import main
        import argparse
        # Build parser the same way main() does internally
        parser = argparse.ArgumentParser()
        parser.add_argument("--xlsx", required=True)
        parser.add_argument("--calendar", required=True, action="append")
        parser.add_argument("--children", required=True)
        args = parser.parse_args([
            "--xlsx", "fake.xlsx",
            "--calendar", "Emma:/path/to/tdsb.md",
            "--calendar", "Liam:/path/to/gist.md",
            "--children", "Emma,Liam",
        ])
        assert args.calendar == ["Emma:/path/to/tdsb.md", "Liam:/path/to/gist.md"]

    def test_single_calendar_via_append(self):
        """Single --calendar should produce a list of one element."""
        import argparse
        parser = argparse.ArgumentParser()
        parser.add_argument("--calendar", required=True, action="append")
        args = parser.parse_args(["--calendar", "/path/to/tdsb.md"])
        assert args.calendar == ["/path/to/tdsb.md"]

    def test_main_calls_multi_when_per_child_calendars(self):
        """When --calendar has 'Child:path' entries, main() should use build_annual_days_multi."""
        tdsb_path = _write_temp_calendar(TDSB_CALENDAR)
        gist_path = _write_temp_calendar(GIST_CALENDAR)
        try:
            # Create a minimal xlsx with Provider Comparison and Daily Schedule
            tmp_xlsx = _create_minimal_xlsx(["Emma", "Liam"])
            with patch("generate_annual_schedule.build_annual_days_multi") as mock_multi, \
                 patch("generate_annual_schedule.build_annual_days") as mock_single:
                mock_multi.return_value = []
                mock_single.return_value = []
                with patch("sys.argv", [
                    "prog", "--xlsx", tmp_xlsx,
                    "--calendar", f"Emma:{tdsb_path}",
                    "--calendar", f"Liam:{gist_path}",
                    "--children", "Emma,Liam",
                    "--pa-day-provider", "TestProvider",
                    "--break-provider", "TestProvider",
                ]):
                    main()
                mock_multi.assert_called_once()
                mock_single.assert_not_called()
        finally:
            os.unlink(tdsb_path)
            os.unlink(gist_path)
            os.unlink(tmp_xlsx)

    def test_main_calls_single_when_one_calendar(self):
        """When --calendar is a plain path, main() should use build_annual_days."""
        tcdsb_path = _write_temp_calendar(TCDSB_CALENDAR)
        try:
            tmp_xlsx = _create_minimal_xlsx(["Emma", "Liam"])
            with patch("generate_annual_schedule.build_annual_days_multi") as mock_multi, \
                 patch("generate_annual_schedule.build_annual_days") as mock_single:
                mock_single.return_value = []
                mock_multi.return_value = []
                with patch("sys.argv", [
                    "prog", "--xlsx", tmp_xlsx,
                    "--calendar", tcdsb_path,
                    "--children", "Emma,Liam",
                    "--pa-day-provider", "TestProvider",
                    "--break-provider", "TestProvider",
                ]):
                    main()
                mock_single.assert_called_once()
                mock_multi.assert_not_called()
        finally:
            os.unlink(tcdsb_path)
            os.unlink(tmp_xlsx)


def _create_minimal_xlsx(children):
    """Create a minimal xlsx with Provider Comparison and Daily Schedule tabs for testing."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = openpyxl.Workbook()

    # Provider Comparison tab
    ws_pc = wb.active
    ws_pc.title = "Provider Comparison"
    ws_pc.cell(row=1, column=1, value="Provider Comparison")
    ws_pc.cell(row=3, column=1, value="Provider")
    ws_pc.cell(row=3, column=3, value="Daily Cost")
    ws_pc.cell(row=4, column=1, value="TestProvider")
    ws_pc.cell(row=4, column=3, value=50)  # daily
    ws_pc.cell(row=4, column=4, value=5)   # before
    ws_pc.cell(row=4, column=5, value=5)   # after
    ws_pc.cell(row=4, column=6, value=2)   # lunch

    # Daily Schedule tab
    ws_ds = wb.create_sheet("Daily Schedule")
    ws_ds.cell(row=3, column=1, value="Date")
    # One summer day
    ws_ds.cell(row=4, column=1, value=date(2025, 6, 30))
    ws_ds.cell(row=4, column=3, value=1)  # week
    offsets = get_child_col_offsets(len(children))
    for i, child in enumerate(children):
        ws_ds.cell(row=4, column=offsets[i] + 1, value="TestProvider")

    wb.save(tmp.name)
    wb.close()
    return tmp.name


def _create_xlsx_with_period_rates():
    """Create xlsx with Provider Comparison that has Summer, PA Day, and Break rate columns."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Provider Comparison"

    # Headers (row 3)
    headers = ["Provider", "Age Range",
               "Daily Cost", "Before Care", "After Care", "Lunch",  # Summer C-F
               "PA Daily", "PA Before", "PA After", "PA Lunch",      # PA Day G-J
               "Brk Daily", "Brk Before", "Brk After", "Brk Lunch"] # Break K-N
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)

    # Provider with all 3 rate sections (row 4)
    ws.cell(row=4, column=1, value="FullRates Camp")
    ws.cell(row=4, column=3, value=60)   # summer daily
    ws.cell(row=4, column=4, value=10)   # summer before
    ws.cell(row=4, column=5, value=10)   # summer after
    ws.cell(row=4, column=6, value=7)    # summer lunch
    ws.cell(row=4, column=7, value=45)   # pa daily
    ws.cell(row=4, column=8, value=8)    # pa before
    ws.cell(row=4, column=9, value=8)    # pa after
    ws.cell(row=4, column=10, value=5)   # pa lunch
    ws.cell(row=4, column=11, value=50)  # break daily
    ws.cell(row=4, column=12, value=8)   # break before
    ws.cell(row=4, column=13, value=8)   # break after
    ws.cell(row=4, column=14, value=5)   # break lunch

    # Provider with only summer rates (row 5) — PA Day and Break cols empty
    ws.cell(row=5, column=1, value="SummerOnly Camp")
    ws.cell(row=5, column=3, value=70)
    ws.cell(row=5, column=4, value=12)
    ws.cell(row=5, column=5, value=12)
    ws.cell(row=5, column=6, value=8)

    wb.save(tmp.name)
    wb.close()
    return tmp.name


class TestReadProviderRatesWithPeriods:
    """Tests for per-period rate reading from xlsx."""

    def test_read_rate_block_with_values(self):
        """_read_rate_block should return a dict when values exist."""
        # Simulate a row of openpyxl cells as simple objects
        from unittest.mock import MagicMock
        row = [MagicMock(value=v) for v in [None, None, 60, 10, 10, 7]]
        result = _read_rate_block(row, 2)
        assert result is not None
        assert result["daily"] == 60
        assert result["total"] == 87

    def test_read_rate_block_all_empty(self):
        """_read_rate_block should return None when all values are empty."""
        from unittest.mock import MagicMock
        row = [MagicMock(value=v) for v in [None, None, None, None, None, None, None, None, None, None]]
        result = _read_rate_block(row, 6)
        assert result is None

    def test_provider_with_all_period_rates(self):
        """Provider with Summer+PA+Break rates should have all three sections."""
        xlsx = _create_xlsx_with_period_rates()
        try:
            rates = read_provider_rates(xlsx)
            camp = rates["FullRates Camp"]
            assert camp["summer"]["total"] == 87
            assert camp["pa_day"]["total"] == 66
            assert camp["break"]["total"] == 71
        finally:
            os.unlink(xlsx)

    def test_provider_with_summer_only(self):
        """Provider with only summer rates should have None for pa_day and break."""
        xlsx = _create_xlsx_with_period_rates()
        try:
            rates = read_provider_rates(xlsx)
            camp = rates["SummerOnly Camp"]
            assert camp["summer"]["total"] == 102
            assert camp["pa_day"] is None
            assert camp["break"] is None
        finally:
            os.unlink(xlsx)

    def test_backward_compat_flat_keys(self):
        """Flat rate keys (daily, before, after, lunch, total) should still work for backward compat."""
        xlsx = _create_xlsx_with_period_rates()
        try:
            rates = read_provider_rates(xlsx)
            camp = rates["FullRates Camp"]
            # Flat keys should be summer rates
            assert camp["total"] == 87
            assert camp["daily"] == 60
        finally:
            os.unlink(xlsx)


class TestRenderMarkdownPeriodRates:
    """Tests that render_markdown uses resolve_period_rate for non-summer periods."""

    def test_pa_day_uses_pa_rate_not_summer(self):
        """PA day cost should use pa_day rate when available, not summer."""
        days = [
            {"date": date(2025, 10, 13), "day_name": "Mon", "period": "pa_day",
             "period_label": "PA Day", "assignments": {"Emma": "FullRates Camp"},
             "notes": "PA Day 1"},
        ]
        rates = {
            "FullRates Camp": {
                "summer": {"daily": 60, "before": 10, "after": 10, "lunch": 7, "total": 87},
                "pa_day": {"daily": 45, "before": 8, "after": 8, "lunch": 5, "total": 66},
                "break": None,
                "daily": 60, "before": 10, "after": 10, "lunch": 7, "total": 87,
            }
        }
        md = render_markdown(days, rates, ["Emma"])
        # Data row should contain $66 (PA day rate)
        table_lines = [l for l in md.split("\n") if "2025-10-13" in l]
        assert len(table_lines) == 1
        assert "$66" in table_lines[0]
        assert "$87" not in table_lines[0]

    def test_winter_break_uses_break_rate(self):
        """Winter break cost should use break rate when available."""
        days = [
            {"date": date(2025, 12, 22), "day_name": "Mon", "period": "winter_break",
             "period_label": "Winter Break", "assignments": {"Emma": "FullRates Camp"},
             "notes": ""},
        ]
        rates = {
            "FullRates Camp": {
                "summer": {"daily": 60, "before": 10, "after": 10, "lunch": 7, "total": 87},
                "pa_day": None,
                "break": {"daily": 50, "before": 8, "after": 8, "lunch": 5, "total": 71},
                "daily": 60, "before": 10, "after": 10, "lunch": 7, "total": 87,
            }
        }
        md = render_markdown(days, rates, ["Emma"])
        table_lines = [l for l in md.split("\n") if "2025-12-22" in l]
        assert len(table_lines) == 1
        assert "$71" in table_lines[0]
        assert "$87" not in table_lines[0]

    def test_summer_still_uses_summer_rate(self):
        """Summer period should continue to use summer rates (no regression)."""
        days = [
            {"date": date(2025, 7, 7), "day_name": "Mon", "period": "summer",
             "period_label": "Summer Wk 1", "assignments": {"Emma": "FullRates Camp"},
             "notes": "Week 1"},
        ]
        rates = {
            "FullRates Camp": {
                "summer": {"daily": 60, "before": 10, "after": 10, "lunch": 7, "total": 87},
                "pa_day": {"daily": 45, "before": 8, "after": 8, "lunch": 5, "total": 66},
                "break": {"daily": 50, "before": 8, "after": 8, "lunch": 5, "total": 71},
                "daily": 60, "before": 10, "after": 10, "lunch": 7, "total": 87,
            }
        }
        md = render_markdown(days, rates, ["Emma"])
        assert "$87" in md

    def test_fallback_to_summer_when_no_period_rate(self):
        """When pa_day rate is None, should fall back to summer rates."""
        days = [
            {"date": date(2025, 10, 13), "day_name": "Mon", "period": "pa_day",
             "period_label": "PA Day", "assignments": {"Emma": "SummerOnly Camp"},
             "notes": "PA Day 1"},
        ]
        rates = {
            "SummerOnly Camp": {
                "summer": {"daily": 70, "before": 12, "after": 12, "lunch": 8, "total": 102},
                "pa_day": None,
                "break": None,
                "daily": 70, "before": 12, "after": 12, "lunch": 8, "total": 102,
            }
        }
        md = render_markdown(days, rates, ["Emma"])
        assert "$102" in md
