"""Excel (xlsx) reading and writing functions for annual schedule generation."""

import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from rate_resolver import _read_rate_block, resolve_period_rate


# ---------------------------------------------------------------------------
# Dynamic child column layout constants
# ---------------------------------------------------------------------------

SHARED_PREFIX_COLS = 3  # Date, Day, Period/Week#
COLS_PER_CHILD = 6      # Camp Name, Before Care, Camp Fee, After Care, Lunch, Day Total


def _require_openpyxl():
    """Lazy-import openpyxl and return (openpyxl, get_column_letter) tuple."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        return openpyxl, get_column_letter
    except ImportError:
        print("Error: openpyxl is required for Excel output. Install with: pip install openpyxl",
              file=sys.stderr)
        sys.exit(1)


def calculate_total_cols(n_children):
    """Total columns = prefix + (N * child_block) + daily_total."""
    return SHARED_PREFIX_COLS + (n_children * COLS_PER_CHILD) + 1


def get_child_col_offsets(n_children):
    """Return 0-indexed column offsets for each child's camp name column."""
    return [SHARED_PREFIX_COLS + i * COLS_PER_CHILD for i in range(n_children)]


def validate_child_count(n):
    if n > 4:
        print(f"Error: Maximum 4 children supported. Got {n}.", file=sys.stderr)
        sys.exit(1)
    if n < 1:
        print("Error: At least 1 child required.", file=sys.stderr)
        sys.exit(1)


def read_provider_rates(xlsx_path):
    """Read provider rates from the Provider Comparison tab.

    Returns {provider: {summer: {daily, before, after, lunch, total},
                        pa_day: {...} or None, break: {...} or None,
                        daily, before, after, lunch, total}}
    Flat keys (daily, before, etc.) are backward-compat aliases for summer rates.

    PA Day (cols G-J) and Break (cols K-N) are only read when the header row
    contains matching labels (e.g., "PA Daily" or "PA Day"). Otherwise these
    columns may hold weekly summaries or other data from older spreadsheets.
    """
    openpyxl, _ = _require_openpyxl()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Provider Comparison"]

    # Detect whether expanded rate columns exist by checking row 3 headers
    header_g = str(ws.cell(row=3, column=7).value or "").lower()
    header_k = str(ws.cell(row=3, column=11).value or "").lower()
    has_pa_cols = "pa" in header_g
    has_break_cols = "br" in header_k  # "Brk Daily" or "Break Daily"

    rates = {}
    for row in ws.iter_rows(min_row=4, max_col=14, values_only=False):
        name = row[0].value  # col A
        if not name or name.startswith("8-Week"):
            break

        # Summer rates (cols C-F, 0-indexed: 2-5)
        summer = _read_rate_block(row, 2)
        if summer is None:
            # No summer rates means this row is not a valid provider
            continue

        # PA Day rates (cols G-J, 0-indexed: 6-9) — only if headers match
        pa_day = _read_rate_block(row, 6) if has_pa_cols else None

        # Break rates (cols K-N, 0-indexed: 10-13) — only if headers match
        brk = _read_rate_block(row, 10) if has_break_cols else None

        rates[name] = {
            "summer": summer,
            "pa_day": pa_day,
            "break": brk,
            # Backward compat: flat rate keys = summer rates
            **summer,
        }
    wb.close()
    return rates


def read_summer_assignments(xlsx_path, children):
    """Read summer camp assignments from the Daily Schedule tab.

    Returns list of dicts: [{date, week, assignments: {child: camp_name}}]
    """
    openpyxl, _ = _require_openpyxl()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Daily Schedule"]
    days = []

    # Child camp columns: dynamically computed based on number of children
    child_cols = get_child_col_offsets(len(children))

    for row in ws.iter_rows(min_row=4, max_col=calculate_total_cols(len(children)), values_only=False):
        date_val = row[0].value  # col A
        if date_val is None or (isinstance(date_val, str) and date_val == "TOTAL"):
            break
        if isinstance(date_val, datetime):
            date_val = date_val.date()
        elif isinstance(date_val, str):
            date_val = datetime.strptime(date_val, "%Y-%m-%d").date()

        week_num = row[2].value  # col C
        assignments = {}
        for i, child in enumerate(children):
            if i < len(child_cols):
                camp = row[child_cols[i]].value or ""
                assignments[child] = camp

        days.append({
            "date": date_val,
            "week": week_num,
            "assignments": assignments,
        })
    wb.close()
    return days


def _vlookup_col_indices(period, has_pa_cols, has_break_cols):
    """Return VLOOKUP column indices (daily, before, after, lunch) based on period type.

    Summer rates: cols C-F (indices 3,4,5,6)
    PA Day rates: cols G-J (indices 7,8,9,10) — only if has_pa_cols
    Break rates:  cols K-N (indices 11,12,13,14) — only if has_break_cols
    Falls back to summer columns when period-specific columns are unavailable.
    """
    if period in ("pa_day", "school_holiday") and has_pa_cols:
        return 7, 8, 9, 10  # daily, before, after, lunch
    elif period in ("winter_break", "march_break", "fall_break") and has_break_cols:
        return 11, 12, 13, 14
    else:
        return 3, 4, 5, 6


def update_xlsx(xlsx_path, annual_days, children, provider_count=3):
    """Add or replace an 'Annual Schedule' tab in the xlsx.

    Uses dynamic column layout: 3 prefix + (N * 6 child block) + 1 daily total.
    VLOOKUP formulas reference Provider Comparison.
    """
    openpyxl, get_column_letter = _require_openpyxl()
    n_children = len(children)
    total_cols = calculate_total_cols(n_children)
    child_offsets = get_child_col_offsets(n_children)

    wb = openpyxl.load_workbook(xlsx_path)

    # Detect whether expanded rate columns exist
    ws_pc = wb["Provider Comparison"]
    header_g = str(ws_pc.cell(row=3, column=7).value or "").lower()
    header_k = str(ws_pc.cell(row=3, column=11).value or "").lower()
    has_pa_cols = "pa" in header_g
    has_break_cols = "br" in header_k

    # Remove existing Annual Schedule tab if present
    if "Annual Schedule" in wb.sheetnames:
        del wb["Annual Schedule"]

    ws = wb.create_sheet("Annual Schedule")

    # VLOOKUP range — extend to col N when period-specific rates exist
    max_vlookup_col = "N" if (has_pa_cols or has_break_cols) else "M"
    vlookup_range = f"'Provider Comparison'!$A$4:${max_vlookup_col}${3 + provider_count}"

    # Row 1: Title
    ws["A1"] = "Annual Schedule"

    # Row 2: Instructions
    ws["A2"] = "Auto-generated from Daily Schedule + school calendar. Camp names reference Provider Comparison tab."

    # Row 3: Headers — dynamically generated for N children
    headers = ["Date", "Day", "Period"]
    for child in children:
        headers.extend([
            f"{child}'s Camp",
            f"{child} Before Care", f"{child} Camp",
            f"{child} After Care", f"{child} Lunch",
            f"{child} Day Total",
        ])
    headers.append("Daily Total")

    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Data rows start at row 4
    daily_total_col = total_cols  # last column

    for i, day in enumerate(annual_days):
        row = i + 4
        d = day["date"]

        # Col A: Date
        ws.cell(row=row, column=1, value=d)
        ws[f"A{row}"].number_format = "YYYY-MM-DD"

        # Col B: Day of week (formula)
        ws.cell(row=row, column=2, value=f'=TEXT(A{row},"ddd")')

        # Col C: Period label
        ws.cell(row=row, column=3, value=day["period_label"])

        # Per-child columns (6 cols each: Camp Name, Before Care, Camp Fee, After Care, Lunch, Day Total)
        # Select VLOOKUP column indices based on period type
        vl_daily, vl_before, vl_after, vl_lunch = _vlookup_col_indices(
            day["period"], has_pa_cols, has_break_cols)

        child_day_total_cols = []
        for ci, child in enumerate(children):
            # 1-indexed column for camp name
            camp_col = child_offsets[ci] + 1  # convert 0-indexed offset to 1-indexed
            camp_letter = get_column_letter(camp_col)

            camp_name = day["assignments"].get(child, "")
            ws.cell(row=row, column=camp_col, value=camp_name)

            # Guard: skip VLOOKUP for empty or "In school" camp names
            guard = f'OR({camp_letter}{row}="",{camp_letter}{row}="In school")'

            # Before Care (offset+1)
            before_col = camp_col + 1
            ws.cell(row=row, column=before_col,
                    value=f'=IF({guard},0,VLOOKUP({camp_letter}{row},{vlookup_range},{vl_before},FALSE))')

            # Camp Fee (offset+2)
            fee_col = camp_col + 2
            ws.cell(row=row, column=fee_col,
                    value=f'=IF({guard},0,VLOOKUP({camp_letter}{row},{vlookup_range},{vl_daily},FALSE))')

            # After Care (offset+3)
            after_col = camp_col + 3
            ws.cell(row=row, column=after_col,
                    value=f'=IF({guard},0,VLOOKUP({camp_letter}{row},{vlookup_range},{vl_after},FALSE))')

            # Lunch (offset+4)
            lunch_col = camp_col + 4
            ws.cell(row=row, column=lunch_col,
                    value=f'=IF({guard},0,VLOOKUP({camp_letter}{row},{vlookup_range},{vl_lunch},FALSE))')

            # Child Day Total (offset+5) = SUM of before..lunch
            day_total_col = camp_col + 5
            before_letter = get_column_letter(before_col)
            lunch_letter = get_column_letter(lunch_col)
            ws.cell(row=row, column=day_total_col,
                    value=f"=SUM({before_letter}{row}:{lunch_letter}{row})")
            child_day_total_cols.append(day_total_col)

        # Daily Total = sum of all child day totals
        if len(child_day_total_cols) == 1:
            dt_letter = get_column_letter(child_day_total_cols[0])
            ws.cell(row=row, column=daily_total_col, value=f"={dt_letter}{row}")
        else:
            parts = "+".join(f"{get_column_letter(c)}{row}" for c in child_day_total_cols)
            ws.cell(row=row, column=daily_total_col, value=f"={parts}")

    # TOTAL row
    total_row = len(annual_days) + 4
    ws.cell(row=total_row, column=1, value="TOTAL")
    last_data_row = total_row - 1

    # SUM formulas for cost columns (all child cost columns + daily total)
    sum_cols = []
    for ci in range(n_children):
        camp_col = child_offsets[ci] + 1
        # Before care, camp fee, after care, lunch, child day total
        sum_cols.extend([camp_col + 1, camp_col + 2, camp_col + 3, camp_col + 4, camp_col + 5])
    sum_cols.append(daily_total_col)

    for col in sum_cols:
        col_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({col_letter}4:{col_letter}{last_data_row})")

    wb.save(xlsx_path)
    wb.close()
